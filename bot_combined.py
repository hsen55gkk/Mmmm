import os
import logging
import sqlite3
from datetime import datetime
from openpyxl import load_workbook, Workbook

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

# --- bot_states.py content ---
class Form(StatesGroup):
    full_name = State()
    dob = State()
    grade = State()
    section = State()
    student_number = State()
    phone_number = State()
    parent_phone_number = State()
    middle_school = State()
    location_link = State()
    address_description = State()
    personal_photo = State()
    student_card_photo = State()
    father_card_photo = State()
    mother_card_photo = State()
    status = State()
    role = State()
    academic_year = State()
    edit_field = State()
    edit_value = State()

class Search(StatesGroup):
    search_name = State()

class AdmissionForm(StatesGroup):
    full_name = State()
    dob = State()
    phone_number = State()
    parent_phone_number = State()
    middle_school = State()
    location_link = State()
    address_description = State()
    personal_photo = State()
    student_card_photo = State()
    father_card_photo = State()
    mother_card_photo = State()
    edit_field = State()
    edit_value = State()

class ContactAdmin(StatesGroup):
    message_text = State()

class FileUpload(StatesGroup):
    waiting_for_file = State()

class Reports(StatesGroup):
    main_menu = State()

class Admin(StatesGroup):
    waiting_for_password = State()
    main_menu = State()
    toggle_admission_form = State()
    manage_admins = State()
    add_admin_id = State()
    remove_admin_id = State()
    allow_deny_student_data_view = State()

# --- End bot_states.py content ---

# --- utils.py content ---
# Database connection (redefined here for combined file)
def get_db_connection():
    conn = sqlite3.connect("students_data.db")
    conn.row_factory = sqlite3.Row
    return conn

async def download_photo(file_id: str, destination_folder: str, bot: Bot) -> str:
    """Downloads a photo from Telegram and saves it to a specified folder."""
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    file_info = await bot.get_file(file_id)
    file_path = file_info.file_path
    
    file_extension = file_path.split(".")[-1]
    destination_path = os.path.join(destination_folder, f"{file_id}.{file_extension}")

    await bot.download_file(file_path, destination_path)
    return destination_path

def process_excel_file(file_path: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Assuming the first column contains full names
    for row in sheet.iter_rows(min_row=1, values_only=True):
        full_name = row[0] # Assuming full name is in the first column
        if not full_name:
            continue

        try:
            # Check if student already exists by full_name
            cursor.execute("SELECT * FROM Students WHERE full_name = ?", (full_name,))
            existing_student = cursor.fetchone()

            if existing_student:
                print(f"Student {full_name} already exists. Skipping insertion.")
                continue

            # Insert only full_name initially, other fields will be null
            cursor.execute("INSERT INTO Students (full_name) VALUES (?) ", (full_name,))
            conn.commit()
        except Exception as e:
            print(f"Error inserting row: {full_name} - {e}")
            conn.rollback()
    conn.close()

def process_word_file(file_path: str):
    doc = Document(file_path)
    conn = get_db_connection()
    cursor = conn.cursor()
    for para in doc.paragraphs:
        full_name = para.text.strip()
        if not full_name:
            continue
        try:
            # Check if student already exists by full_name
            cursor.execute("SELECT * FROM Students WHERE full_name = ?", (full_name,))
            existing_student = cursor.fetchone()

            if existing_student:
                print(f"Student {full_name} already exists. Skipping insertion.")
                continue

            # Insert only full_name initially, other fields will be null
            cursor.execute("INSERT INTO Students (full_name) VALUES (?) ", (full_name,))
            conn.commit()
        except Exception as e:
            print(f"Error inserting row: {full_name} - {e}")
            conn.rollback()
    conn.close()

def get_student_statistics():
    conn = get_db_connection()
    cursor = conn.cursor()
    stats = {}

    # Total students
    cursor.execute("SELECT COUNT(*) FROM Students")
    stats["total_students"] = cursor.fetchone()[0]

    # Students by grade
    cursor.execute("SELECT grade, COUNT(*) FROM Students GROUP BY grade")
    stats["students_by_grade"] = dict(cursor.fetchall())

    # Students by section
    cursor.execute("SELECT section, COUNT(*) FROM Students GROUP BY section")
    stats["students_by_section"] = dict(cursor.fetchall())

    conn.close()
    return stats


def export_students_to_excel(file_path: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM Students")
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        return False

    workbook = Workbook()
    sheet = workbook.active

    # Write headers
    headers = [description[0] for description in cursor.description]
    sheet.append(headers)

    # Write data
    for row in rows:
        sheet.append(list(row))

    workbook.save(file_path)
    return True

def add_supervisor(telegram_id: int, username: str, full_name: str, password: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Supervisors (telegram_id, username, full_name, password) VALUES (?, ?, ?, ?)",
                       (telegram_id, username, full_name, password))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False # Supervisor with this telegram_id already exists
    finally:
        conn.close()

def remove_supervisor(telegram_id: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Supervisors WHERE telegram_id = ?", (telegram_id,))
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()

def get_all_supervisors():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT telegram_id, username, full_name FROM Supervisors")
    supervisors = cursor.fetchall()
    conn.close()
    return supervisors

def is_supervisor(telegram_id: int, password: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM Supervisors WHERE telegram_id = ? AND password = ?", (telegram_id, password))
    supervisor = cursor.fetchone()
    conn.close()
    return supervisor is not None

# --- End utils.py content ---

# --- create_db.py content (integrated as a function) ---
def create_database():
    conn = sqlite3.connect("students_data.db")
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE,
            full_name TEXT NOT NULL,
            dob TEXT,
            grade TEXT,
            section TEXT,
            student_number INTEGER UNIQUE,
            phone_number TEXT,
            parent_phone_number TEXT,
            middle_school TEXT,
            location_link TEXT,
            address_description TEXT,
            personal_photo_path TEXT,
            student_card_photo_path TEXT,
            father_card_photo_path TEXT,
            mother_card_photo_path TEXT,
            status TEXT,
            role TEXT,
            academic_year TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            is_form_locked BOOLEAN DEFAULT FALSE,
            can_view_data BOOLEAN DEFAULT TRUE
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Admission_Requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER,
            full_name TEXT NOT NULL,
            dob TEXT,
            phone_number TEXT,
            parent_phone_number TEXT,
            middle_school TEXT,
            location_link TEXT,
            address_description TEXT,
            personal_photo_path TEXT,
            student_card_photo_path TEXT,
            father_card_photo_path TEXT,
            mother_card_photo_path TEXT,
            status TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            setting_name TEXT UNIQUE NOT NULL,
            setting_value TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Supervisors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL,
            username TEXT,
            full_name TEXT,
            password TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Insert default settings if they don\'t exist
    cursor.execute("INSERT OR IGNORE INTO Settings (setting_name, setting_value) VALUES (?, ?)", ("form_status", "open"))

    conn.commit()
    conn.close()
# --- End create_db.py content ---

# Configure logging
logging.basicConfig(level=logging.INFO)

# Admin Telegram ID (replace with actual admin ID)
ADMIN_TELEGRAM_ID = 1738750806 # TODO: Replace with actual admin Telegram ID
ADMIN_PASSWORD = "1526374850"

# Helper function to get setting from DB
def get_setting(setting_name):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT setting_value FROM Settings WHERE setting_name = ?", (setting_name,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None

# Helper function to update setting in DB
def update_setting(setting_name, setting_value):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE Settings SET setting_value = ? WHERE setting_name = ?", (setting_value, setting_name))
    conn.commit()
    conn.close()

async def main() -> None:
    # Ensure database is created before starting the bot
    create_database()

    BOT_TOKEN = os.getenv("BOT_TOKEN")
    if not BOT_TOKEN:
        print("Error: BOT_TOKEN environment variable is not set. Please set it to run the bot.")
        return

    bot = Bot(token=BOT_TOKEN)
    dp = Dispatcher()

    # Register handlers here after dp is initialized
    @dp.message(CommandStart())
    async def command_start_handler(message: types.Message) -> None:
        """This handler receives messages with `/start` command"""
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="تسجيل طالب جديد")],
                [types.KeyboardButton(text="البحث عن اسمي")],
                [types.KeyboardButton(text="طلب تقديم إلى إعدادية المنتظر للبنين")],
                [types.KeyboardButton(text="تواصل مع الإدارة")],
                [types.KeyboardButton(text="رفع ملف بيانات")],
                [types.KeyboardButton(text="مشرف")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer(f"مرحباً بك يا {message.from_user.full_name}! أنا بوت إدارة بيانات الطلاب لإعدادية المنتظر للبنين. كيف يمكنني مساعدتك اليوم؟", reply_markup=keyboard)

    @dp.message(F.text == "تسجيل طالب جديد")
    async def cmd_register_student(message: types.Message, state: FSMContext):
        form_status = get_setting("form_status")
        if form_status == "closed":
            await message.answer("عذراً، استمارة التسجيل مغلقة حالياً.")
            await state.clear()
            return
        await state.set_state(Form.full_name)
        await message.answer("أهلاً بك في استمارة تسجيل الطلاب. يرجى إدخال الاسم الرباعي:")

    @dp.message(Form.full_name)
    async def process_full_name(message: types.Message, state: FSMContext):
        await state.update_data(full_name=message.text)
        await state.set_state(Form.dob)
        await message.answer("يرجى إدخال تاريخ الميلاد (مثال: 2005-01-15):")

    @dp.message(Form.dob)
    async def process_dob(message: types.Message, state: FSMContext):
        try:
            datetime.strptime(message.text, "%Y-%m-%d")
            await state.update_data(dob=message.text)
            await state.set_state(Form.grade)
            await message.answer("يرجى إدخال الصف (الرابع، الخامس، السادس):")
        except ValueError:
            await message.answer("صيغة تاريخ الميلاد غير صحيحة. يرجى استخدام الصيغة YYYY-MM-DD (مثال: 2005-01-15):")

    @dp.message(Form.grade)
    async def process_grade(message: types.Message, state: FSMContext):
        if message.text in ["الرابع", "الخامس", "السادس"]:
            await state.update_data(grade=message.text)
            await state.set_state(Form.section)
            await message.answer("يرجى إدخال الشعبة (أ، ب، ج، د، هـ):")
        else:
            await message.answer("الصف غير صحيح. يرجى الاختيار من (الرابع، الخامس، السادس):")

    @dp.message(Form.section)
    async def process_section(message: types.Message, state: FSMContext):
        if message.text in ["أ", "ب", "ج", "د", "هـ"]:
            await state.update_data(section=message.text)
            await state.set_state(Form.student_number)
            await message.answer("يرجى إدخال الرقم (من 1 إلى 1000):")
        else:
            await message.answer("الشعبة غير صحيحة. يرجى الاختيار من (أ، ب، ج، د، هـ):")

    @dp.message(Form.student_number)
    async def process_student_number(message: types.Message, state: FSMContext):
        try:
            student_num = int(message.text)
            if 1 <= student_num <= 1000:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM Students WHERE student_number = ?", (student_num,))
                if cursor.fetchone():
                    await message.answer("هذا الرقم مستخدم بالفعل. يرجى إدخال رقم آخر:")
                else:
                    await state.update_data(student_number=student_num)
                    await state.set_state(Form.phone_number)
                    await message.answer("يرجى إدخال رقم هاتف الطالب:")
                conn.close()
            else:
                await message.answer("الرقم يجب أن يكون بين 1 و 1000. يرجى إدخال رقم صحيح:")
        except ValueError:
            await message.answer("الرقم غير صحيح. يرجى إدخال رقم:")

    @dp.message(Form.phone_number)
    async def process_phone_number(message: types.Message, state: FSMContext):
        await state.update_data(phone_number=message.text)
        await state.set_state(Form.parent_phone_number)
        await message.answer("يرجى إدخال رقم هاتف ولي الأمر:")

    @dp.message(Form.parent_phone_number)
    async def process_parent_phone_number(message: types.Message, state: FSMContext):
        await state.update_data(parent_phone_number=message.text)
        await state.set_state(Form.middle_school)
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="متوسطة المجتبى")],
                [types.KeyboardButton(text="متوسطة الصناديد")],
                [types.KeyboardButton(text="أخرى")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer("يرجى اختيار المدرسة المتوسطة التي تخرج منها الطالب:", reply_markup=keyboard)

    @dp.message(Form.middle_school)
    async def process_middle_school(message: types.Message, state: FSMContext):
        if message.text == "أخرى":
            await state.set_state(Form.middle_school)
            await message.answer("يرجى إدخال اسم المدرسة المتوسطة:")
        else:
            await state.update_data(middle_school=message.text)
            await state.set_state(Form.location_link)
            await message.answer("يرجى إرسال رابط الموقع الجغرافي أو اكتب \"لا يوجد\":")

    @dp.message(Form.location_link)
    async def process_location_link(message: types.Message, state: FSMContext):
        await state.update_data(location_link=message.text)
        await state.set_state(Form.address_description)
        await message.answer("يرجى إدخال وصف السكن:")

    @dp.message(Form.address_description)
    async def process_address_description(message: types.Message, state: FSMContext):
        await state.update_data(address_description=message.text)
        await state.set_state(Form.personal_photo)
        await message.answer("يرجى إرسال صورة شخصية واضحة (مباشرة من الكاميرا أو من المعرض):")

    @dp.message(Form.personal_photo, F.photo)
    async def process_personal_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/personal", bot)
        if photo_path:
            await state.update_data(personal_photo_path=photo_path)
            await state.set_state(Form.student_card_photo)
            await message.answer("يرجى إرسال صورة بطاقة الطالب:")
        else:
            await message.answer("حدث خطأ أثناء تحميل الصورة الشخصية. يرجى المحاولة مرة أخرى.")

    @dp.message(Form.student_card_photo, F.photo)
    async def process_student_card_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/student_cards", bot)
        if photo_path:
            await state.update_data(student_card_photo_path=photo_path)
            await state.set_state(Form.father_card_photo)
            await message.answer("يرجى إرسال صورة بطاقة الأب:")
        else:
            await message.answer("حدث خطأ أثناء تحميل صورة بطاقة الطالب. يرجى المحاولة مرة أخرى.")

    @dp.message(Form.father_card_photo, F.photo)
    async def process_father_card_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/father_cards", bot)
        if photo_path:
            await state.update_data(father_card_photo_path=photo_path)
            await state.set_state(Form.mother_card_photo)
            await message.answer("يرجى إرسال صورة بطاقة الأم:")
        else:
            await message.answer("حدث خطأ أثناء تحميل صورة بطاقة الأب. يرجى المحاولة مرة أخرى.")

    @dp.message(Form.mother_card_photo, F.photo)
    async def process_mother_card_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/mother_cards", bot)
        if photo_path:
            await state.update_data(mother_card_photo_path=photo_path)
            await state.set_state(Form.status)
            keyboard = types.ReplyKeyboardMarkup(
                keyboard=[
                    [types.KeyboardButton(text="ناجح")],
                    [types.KeyboardButton(text="راسب")],
                    [types.KeyboardButton(text="مكمل")]
                ],
                resize_keyboard=True,
                one_time_keyboard=True
            )
            await message.answer("يرجى تحديد حالة الطالب:", reply_markup=keyboard)
        else:
            await message.answer("حدث خطأ أثناء تحميل صورة بطاقة الأم. يرجى المحاولة مرة أخرى.")

    @dp.message(Form.status)
    async def process_status(message: types.Message, state: FSMContext):
        if message.text in ["ناجح", "راسب", "مكمل"]:
            await state.update_data(status=message.text)
            await state.set_state(Form.role)
            keyboard = types.ReplyKeyboardMarkup(
                keyboard=[
                    [types.KeyboardButton(text="أول")],
                    [types.KeyboardButton(text="ثاني")],
                    [types.KeyboardButton(text="ثالث")]
                ],
                resize_keyboard=True,
                one_time_keyboard=True
            )
            await message.answer("يرجى تحديد الدور:", reply_markup=keyboard)
        else:
            await message.answer("الدور غير صحيح. يرجى الاختيار من (أول، ثاني، ثالث):")

    @dp.message(Form.role)
    async def process_role(message: types.Message, state: FSMContext):
        if message.text in ["أول", "ثاني", "ثالث"]:
            await state.update_data(role=message.text)
            await state.set_state(Form.academic_year)
            await message.answer("يرجى إدخال العام الدراسي (مثال: 2024-2025):")
        else:
            await message.answer("الدور غير صحيح. يرجى الاختيار من (أول، ثاني، ثالث):")

    @dp.message(Form.academic_year)
    async def process_academic_year(message: types.Message, state: FSMContext):
        await state.update_data(academic_year=message.text)
        user_data = await state.get_data()

        # Display collected data for review
        review_message = "\n".join([
            f'الاسم الرباعي: {user_data.get("full_name")}',
            f'تاريخ الميلاد: {user_data.get("dob")}',
            f'الصف: {user_data.get("grade")}',
            f'الشعبة: {user_data.get("section")}',
            f'الرقم: {user_data.get("student_number")}',
            f'رقم الهاتف: {user_data.get("phone_number")}',
            f'رقم هاتف ولي الأمر: {user_data.get("parent_phone_number")}',
            f'المدرسة المتوسطة: {user_data.get("middle_school")}',
            f'رابط الموقع الجغرافي: {user_data.get("location_link")}',
            f'وصف السكن: {user_data.get("address_description")}',
            f'الحالة: {user_data.get("status")}',
            f'الدور: {user_data.get("role")}',
            f'العام الدراسي: {user_data.get("academic_year")}'
        ])

        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [types.InlineKeyboardButton(text="تعديل قبل الإرسال", callback_data="edit_form")],
                [types.InlineKeyboardButton(text="تأكيد وإرسال", callback_data="submit_form")]
            ]
        )
        await message.answer(f"يرجى مراجعة بياناتك:\n{review_message}", reply_markup=keyboard)
        await state.set_state(Form.academic_year) # Keep state for review/edit

    @dp.callback_query(F.data == "submit_form")
    async def submit_form(callback_query: types.CallbackQuery, state: FSMContext):
        user_data = await state.get_data()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            # Check if student exists by full_name and telegram_id is null
            cursor.execute("SELECT id FROM Students WHERE full_name = ? AND telegram_id IS NULL", (user_data.get("full_name"),))
            student_id_row = cursor.fetchone()

            if student_id_row:
                student_id = student_id_row[0]
                cursor.execute("""
                    UPDATE Students SET
                        telegram_id = ?, dob = ?, grade = ?, section = ?, student_number = ?,
                        phone_number = ?, parent_phone_number = ?, middle_school = ?, location_link = ?,
                        address_description = ?, personal_photo_path = ?, student_card_photo_path = ?,
                        father_card_photo_path = ?, mother_card_photo_path = ?, status = ?, role = ?,
                        academic_year = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (
                    callback_query.from_user.id, user_data.get("dob"), user_data.get("grade"),
                    user_data.get("section"), user_data.get("student_number"), user_data.get("phone_number"),
                    user_data.get("parent_phone_number"), user_data.get("middle_school"),
                    user_data.get("location_link"), user_data.get("address_description"),
                    user_data.get("personal_photo_path"), user_data.get("student_card_photo_path"),
                    user_data.get("father_card_photo_path"), user_data.get("mother_card_photo_path"),
                    user_data.get("status"), user_data.get("role"), user_data.get("academic_year"),
                    student_id
                ))
                await callback_query.message.answer("تم تحديث بياناتك بنجاح!")
            else:
                # If student does not exist or telegram_id is already set, insert as new
                cursor.execute("""
                    INSERT INTO Students (
                        telegram_id, full_name, dob, grade, section, student_number, phone_number,
                        parent_phone_number, middle_school, location_link, address_description,
                        personal_photo_path, student_card_photo_path, father_card_photo_path,
                        mother_card_photo_path, status, role, academic_year
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    callback_query.from_user.id, user_data.get("full_name"), user_data.get("dob"),
                    user_data.get("grade"), user_data.get("section"), user_data.get("student_number"),
                    user_data.get("phone_number"), user_data.get("parent_phone_number"),
                    user_data.get("middle_school"), user_data.get("location_link"),
                    user_data.get("address_description"), user_data.get("personal_photo_path"),
                    user_data.get("student_card_photo_path"), user_data.get("father_card_photo_path"),
                    user_data.get("mother_card_photo_path"), user_data.get("status"),
                    user_data.get("role"), user_data.get("academic_year")
                ))
                await callback_query.message.answer("تم حفظ بياناتك بنجاح! سيتم مراجعتها من قبل الإدارة.")
            conn.commit()
            await state.clear()
        except sqlite3.IntegrityError as e:
            await callback_query.message.answer(f"حدث خطأ أثناء حفظ البيانات: {e}. يرجى المحاولة مرة أخرى.")
        finally:
            conn.close()
        await callback_query.answer()

    @dp.callback_query(F.data == "edit_form")
    async def edit_form(callback_query: types.CallbackQuery, state: FSMContext):
        await callback_query.message.answer("ما هو الحقل الذي تود تعديله؟ (مثال: الاسم الرباعي، تاريخ الميلاد، الصف)")
        await state.set_state(Form.edit_field) # A new state for editing

    @dp.message(Form.edit_field)
    async def process_edit_field(message: types.Message, state: FSMContext):
        field_name = message.text.strip()
        user_data = await state.get_data()
        if field_name in user_data:
            await state.update_data(field_to_edit=field_name)
            await state.set_state(Form.edit_value)
            await message.answer(f"يرجى إدخال القيمة الجديدة للحقل \"{field_name}\":")
        else:
            await message.answer("هذا الحقل غير موجود أو لا يمكن تعديله حاليًا. يرجى إدخال اسم حقل صحيح.")

    @dp.message(Form.edit_value)
    async def process_edit_value(message: types.Message, state: FSMContext):
        user_data = await state.get_data()
        field_to_edit = user_data.get("field_to_edit")
        new_value = message.text

        await state.update_data(**{field_to_edit: new_value})
        updated_data = await state.get_data()

        review_message = "\n".join([
            f'الاسم الرباعي: {updated_data.get("full_name")}',
            f'تاريخ الميلاد: {updated_data.get("dob")}',
            f'الصف: {updated_data.get("grade")}',
            f'الشعبة: {updated_data.get("section")}',
            f'الرقم: {updated_data.get("student_number")}',
            f'رقم الهاتف: {updated_data.get("phone_number")}',
            f'رقم هاتف ولي الأمر: {updated_data.get("parent_phone_number")}',
            f'المدرسة المتوسطة: {updated_data.get("middle_school")}',
            f'رابط الموقع الجغرافي: {updated_data.get("location_link")}',
            f'وصف السكن: {updated_data.get("address_description")}',
            f'الحالة: {updated_data.get("status")}',
            f'الدور: {updated_data.get("role")}',
            f'العام الدراسي: {updated_data.get("academic_year")}'
        ])

        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [types.InlineKeyboardButton(text="تعديل قبل الإرسال", callback_data="edit_form")],
                [types.InlineKeyboardButton(text="تأكيد وإرسال", callback_data="submit_form")]
            ]
        )
        await message.answer(f"تم تحديث الحقل. يرجى مراجعة بياناتك مرة أخرى:\n{review_message}", reply_markup=keyboard)
        await state.set_state(Form.academic_year) # Return to review state

    # Search student functionality
    @dp.message(F.text == "البحث عن اسمي")
    async def cmd_search_student(message: types.Message, state: FSMContext):
        await state.set_state(Search.search_name)
        await message.answer("يرجى إدخال الاسم الرباعي للبحث عنه:")

    @dp.message(Search.search_name)
    async def process_search_name(message: types.Message, state: FSMContext):
        full_name = message.text.strip()
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Students WHERE full_name = ?", (full_name,))
        student = cursor.fetchone()
        conn.close()

        if student:
            student_data = dict(student)
            # Check if telegram_id is null or if any required fields are missing
            if student_data.get("telegram_id") is None or \
               any(student_data.get(field) is None for field in [
                   "dob", "grade", "section", "student_number", "phone_number", "parent_phone_number",
                   "middle_school", "location_link", "address_description", "personal_photo_path",
                   "student_card_photo_path", "father_card_photo_path", "mother_card_photo_path",
                   "status", "role", "academic_year"
               ]):
                # Student found but data is incomplete or telegram_id is null, offer to complete
                await state.set_data(student_data) # Load existing data into FSM context
                await state.set_state(Form.dob) # Start from dob to complete the form
                await message.answer(f"تم العثور على اسمك: {full_name}. يرجى استكمال بياناتك. يرجى إدخال تاريخ الميلاد (مثال: 2005-01-15):")
            else:
                # Student found and data is complete, display it
                if student_data.get("can_view_data") == 0: # Check if can_view_data is FALSE (0)
                    await message.answer("عذراً، لا يمكنك عرض بياناتك حالياً. يرجى التواصل مع الإدارة.")
                    await state.clear()
                    return

                review_message = "\n".join([
                    f'الاسم الرباعي: {student_data.get("full_name")}',
                    f'تاريخ الميلاد: {student_data.get("dob")}',
                    f'الصف: {student_data.get("grade")}',
                    f'الشعبة: {student_data.get("section")}',
                    f'الرقم: {student_data.get("student_number")}',
                    f'رقم الهاتف: {student_data.get("phone_number")}',
                    f'رقم هاتف ولي الأمر: {student_data.get("parent_phone_number")}',
                    f'المدرسة المتوسطة: {student_data.get("middle_school")}',
                    f'رابط الموقع الجغرافي: {student_data.get("location_link")}',
                    f'وصف السكن: {student_data.get("address_description")}',
                    f'الحالة: {student_data.get("status")}',
                    f'الدور: {student_data.get("role")}',
                    f'العام الدراسي: {student_data.get("academic_year")}'
                ])
                keyboard = types.InlineKeyboardMarkup(
                    inline_keyboard=[
                        [types.InlineKeyboardButton(text="تعديل بياناتي", callback_data=f'update_student_{student_data.get("telegram_id")}')]
                    ]
                )
                await message.answer(f"تم العثور على بياناتك:\n{review_message}", reply_markup=keyboard)
                await state.clear()
        else:
            keyboard = types.InlineKeyboardMarkup(
                inline_keyboard=[
                    [types.InlineKeyboardButton(text="تسجيل جديد", callback_data="register_new_student")]
                ]
            )
            await message.answer("لم يتم العثور على اسمك. هل ترغب في تسجيل جديد؟", reply_markup=keyboard)
            await state.clear()

    @dp.callback_query(F.data.startswith("update_student_"))
    async def update_student_data(callback_query: types.CallbackQuery, state: FSMContext):
        telegram_id = int(callback_query.data.split("_")[2])
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Students WHERE telegram_id = ?", (telegram_id,))
        student_data = cursor.fetchone()
        conn.close()

        if student_data:
            await state.set_data(dict(student_data)) # Load existing data into FSM context
            await state.set_state(Form.edit_field)
            await callback_query.message.answer("ما هو الحقل الذي تود تعديله؟ (مثال: رقم الهاتف، وصف السكن)")
        else:
            await callback_query.message.answer("لم يتم العثور على بيانات الطالب.")
        await callback_query.answer()

    @dp.callback_query(F.data == "register_new_student")
    async def register_new_student_callback(callback_query: types.CallbackQuery, state: FSMContext):
        await state.set_state(Form.full_name)
        await callback_query.message.answer("أهلاً بك في استمارة تسجيل الطلاب. يرجى إدخال الاسم الرباعي:")
        await callback_query.answer()

    # Admission Form Handlers
    @dp.message(F.text == "طلب تقديم إلى إعدادية المنتظر للبنين")
    async def cmd_admission_form(message: types.Message, state: FSMContext):
        form_status = get_setting("form_status")
        if form_status == "closed":
            await message.answer("عذراً، استمارة التقديم مغلقة حالياً.")
            await state.clear()
            return
        await state.set_state(AdmissionForm.full_name)
        await message.answer("أهلاً بك في استمارة طلب التقديم. يرجى إدخال الاسم الرباعي:")

    @dp.message(AdmissionForm.full_name)
    async def process_admission_full_name(message: types.Message, state: FSMContext):
        await state.update_data(full_name=message.text)
        await state.set_state(AdmissionForm.dob)
        await message.answer("يرجى إدخال تاريخ الميلاد (مثال: 2005-01-15):")

    @dp.message(AdmissionForm.dob)
    async def process_admission_dob(message: types.Message, state: FSMContext):
        try:
            datetime.strptime(message.text, "%Y-%m-%d")
            await state.update_data(dob=message.text)
            await state.set_state(AdmissionForm.phone_number)
            await message.answer("يرجى إدخال رقم هاتف الطالب:")
        except ValueError:
            await message.answer("صيغة تاريخ الميلاد غير صحيحة. يرجى استخدام الصيغة YYYY-MM-DD (مثال: 2005-01-15):")

    @dp.message(AdmissionForm.phone_number)
    async def process_admission_phone_number(message: types.Message, state: FSMContext):
        await state.update_data(phone_number=message.text)
        await state.set_state(AdmissionForm.parent_phone_number)
        await message.answer("يرجى إدخال رقم هاتف ولي الأمر:")

    @dp.message(AdmissionForm.parent_phone_number)
    async def process_admission_parent_phone_number(message: types.Message, state: FSMContext):
        await state.update_data(parent_phone_number=message.text)
        await state.set_state(AdmissionForm.middle_school)
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="متوسطة المجتبى")],
                [types.KeyboardButton(text="متوسطة الصناديد")],
                [types.KeyboardButton(text="أخرى")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer("يرجى اختيار المدرسة المتوسطة التي تخرج منها الطالب:", reply_markup=keyboard)

    @dp.message(AdmissionForm.middle_school)
    async def process_admission_middle_school(message: types.Message, state: FSMContext):
        if message.text == "أخرى":
            await state.set_state(AdmissionForm.middle_school)
            await message.answer("يرجى إدخال اسم المدرسة المتوسطة:")
        else:
            await state.update_data(middle_school=message.text)
            await state.set_state(AdmissionForm.location_link)
            await message.answer("يرجى إرسال رابط الموقع الجغرافي أو اكتب \"لا يوجد\":")

    @dp.message(AdmissionForm.location_link)
    async def process_admission_location_link(message: types.Message, state: FSMContext):
        await state.update_data(location_link=message.text)
        await state.set_state(AdmissionForm.address_description)
        await message.answer("يرجى إدخال وصف السكن:")

    @dp.message(AdmissionForm.address_description)
    async def process_admission_address_description(message: types.Message, state: FSMContext):
        await state.update_data(address_description=message.text)
        await state.set_state(AdmissionForm.personal_photo)
        await message.answer("يرجى إرسال صورة شخصية واضحة (مباشرة من الكاميرا أو من المعرض):")

    @dp.message(AdmissionForm.personal_photo, F.photo)
    async def process_admission_personal_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/admission_personal", bot)
        if photo_path:
            await state.update_data(personal_photo_path=photo_path)
            await state.set_state(AdmissionForm.student_card_photo)
            await message.answer("يرجى إرسال صورة بطاقة الطالب:")
        else:
            await message.answer("حدث خطأ أثناء تحميل الصورة الشخصية. يرجى المحاولة مرة أخرى.")

    @dp.message(AdmissionForm.student_card_photo, F.photo)
    async def process_admission_student_card_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/admission_student_cards", bot)
        if photo_path:
            await state.update_data(student_card_photo_path=photo_path)
            await state.set_state(AdmissionForm.father_card_photo)
            await message.answer("يرجى إرسال صورة بطاقة الأب:")
        else:
            await message.answer("حدث خطأ أثناء تحميل صورة بطاقة الطالب. يرجى المحاولة مرة أخرى.")

    @dp.message(AdmissionForm.father_card_photo, F.photo)
    async def process_admission_father_card_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/father_cards", bot)
        if photo_path:
            await state.update_data(father_card_photo_path=photo_path)
            await state.set_state(AdmissionForm.mother_card_photo)
            await message.answer("يرجى إرسال صورة بطاقة الأم:")
        else:
            await message.answer("حدث خطأ أثناء تحميل صورة بطاقة الأب. يرجى المحاولة مرة أخرى.")

    @dp.message(AdmissionForm.mother_card_photo, F.photo)
    async def process_admission_mother_card_photo(message: types.Message, state: FSMContext):
        file_id = message.photo[-1].file_id
        photo_path = await download_photo(file_id, "photos/admission_mother_cards", bot)
        if photo_path:
            await state.update_data(mother_card_photo_path=photo_path)
            user_data = await state.get_data()

            review_message = "\n".join([
                f'الاسم الرباعي: {user_data.get("full_name")}',
                f'تاريخ الميلاد: {user_data.get("dob")}',
                f'رقم الهاتف: {user_data.get("phone_number")}',
                f'رقم هاتف ولي الأمر: {user_data.get("parent_phone_number")}',
                f'المدرسة المتوسطة: {user_data.get("middle_school")}',
                f'رابط الموقع الجغرافي: {user_data.get("location_link")}',
                f'وصف السكن: {user_data.get("address_description")}'
            ])

            keyboard = types.InlineKeyboardMarkup(
                inline_keyboard=[
                    [types.InlineKeyboardButton(text="تعديل قبل الإرسال", callback_data="edit_admission_form")],
                    [types.InlineKeyboardButton(text="تأكيد وإرسال", callback_data="submit_admission_form")]
                ]
            )
            await message.answer(f"يرجى مراجعة بيانات طلب التقديم:\n{review_message}", reply_markup=keyboard)
            await state.set_state(AdmissionForm.mother_card_photo) # Keep state for review/edit
        else:
            await message.answer("حدث خطأ أثناء تحميل صورة بطاقة الأم. يرجى المحاولة مرة أخرى.")

    @dp.callback_query(F.data == "submit_admission_form")
    async def submit_admission_form(callback_query: types.CallbackQuery, state: FSMContext):
        user_data = await state.get_data()
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO Admission_Requests (
                    telegram_id, full_name, dob, phone_number, parent_phone_number,
                    middle_school, location_link, address_description, personal_photo_path,
                    student_card_photo_path, father_card_photo_path, mother_card_photo_path, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                callback_query.from_user.id, user_data.get("full_name"), user_data.get("dob"),
                user_data.get("phone_number"), user_data.get("parent_phone_number"),
                user_data.get("middle_school"), user_data.get("location_link"),
                user_data.get("address_description"), user_data.get("personal_photo_path"),
                user_data.get("student_card_photo_path"), user_data.get("father_card_photo_path"),
                user_data.get("mother_card_photo_path"), "قيد المراجعة"
            ))
            conn.commit()
            await callback_query.message.answer("تم إرسال طلب التقديم بنجاح! سيتم مراجعته من قبل الإدارة.")
            await state.clear()
        except sqlite3.IntegrityError as e:
            await callback_query.message.answer(f"حدث خطأ أثناء حفظ البيانات: {e}. يرجى المحاولة مرة أخرى.")
        finally:
            conn.close()
        await callback_query.answer()

    @dp.callback_query(F.data == "edit_admission_form")
    async def edit_admission_form(callback_query: types.CallbackQuery, state: FSMContext):
        await callback_query.message.answer("ما هو الحقل الذي تود تعديله في استمارة التقديم؟ (مثال: الاسم الرباعي، تاريخ الميلاد)")
        await state.set_state(AdmissionForm.edit_field) # A new state for editing admission form

    @dp.message(AdmissionForm.edit_field)
    async def process_admission_edit_field(message: types.Message, state: FSMContext):
        field_name = message.text.strip()
        user_data = await state.get_data()
        if field_name in user_data:
            await state.update_data(field_to_edit=field_name)
            await state.set_state(AdmissionForm.edit_value)
            await message.answer(f"يرجى إدخال القيمة الجديدة للحقل \"{field_name}\":")
        else:
            await message.answer("هذا الحقل غير موجود أو لا يمكن تعديله حاليًا. يرجى إدخال اسم حقل صحيح.")

    @dp.message(AdmissionForm.edit_value)
    async def process_admission_edit_value(message: types.Message, state: FSMContext):
        user_data = await state.get_data()
        field_to_edit = user_data.get("field_to_edit")
        new_value = message.text

        await state.update_data(**{field_to_edit: new_value})
        updated_data = await state.get_data()

        review_message = "\n".join([
            f'الاسم الرباعي: {updated_data.get("full_name")}',
            f'تاريخ الميلاد: {updated_data.get("dob")}',
            f'رقم الهاتف: {updated_data.get("phone_number")}',
            f'رقم هاتف ولي الأمر: {updated_data.get("parent_phone_number")}',
            f'المدرسة المتوسطة: {updated_data.get("middle_school")}',
            f'رابط الموقع الجغرافي: {updated_data.get("location_link")}',
            f'وصف السكن: {updated_data.get("address_description")}'
        ])

        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [types.InlineKeyboardButton(text="تعديل قبل الإرسال", callback_data="edit_admission_form")],
                [types.InlineKeyboardButton(text="تأكيد وإرسال", callback_data="submit_admission_form")]
            ]
        )
        await message.answer(f"تم تحديث الحقل. يرجى مراجعة بياناتك مرة أخرى:\n{review_message}", reply_markup=keyboard)
        await state.set_state(AdmissionForm.mother_card_photo) # Return to review state

    # Contact Admin Handlers
    @dp.message(F.text == "تواصل مع الإدارة")
    async def cmd_contact_admin(message: types.Message, state: FSMContext):
        await state.set_state(ContactAdmin.message_text)
        await message.answer("يرجى كتابة رسالتك للإدارة:")

    @dp.message(ContactAdmin.message_text)
    async def process_contact_admin_message(message: types.Message, state: FSMContext):
        user_message = message.text
        user_info = f"From: {message.from_user.full_name} (ID: {message.from_user.id})\n"
        full_message = f"{user_info}Message: {user_message}"

        try:
            await bot.send_message(ADMIN_TELEGRAM_ID, full_message)
            await message.answer("تم إرسال رسالتك إلى الإدارة بنجاح.")
        except Exception as e:
            await message.answer(f"حدث خطأ أثناء إرسال رسالتك: {e}")
        finally:
            await state.clear()

    # File Upload Handlers
    @dp.message(F.text == "رفع ملف بيانات")
    async def cmd_upload_file(message: types.Message, state: FSMContext):
        await state.set_state(FileUpload.waiting_for_file)
        await message.answer("يرجى إرسال ملف Excel (.xlsx) أو Word (.docx) الذي يحتوي على بيانات الطلاب.")

    @dp.message(FileUpload.waiting_for_file, F.document)
    async def process_uploaded_file(message: types.Message, state: FSMContext):
        file_id = message.document.file_id
        file_name = message.document.file_name
        file_path = f"downloads/{file_name}"

        # Ensure the downloads directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        await bot.download_file_by_id(file_id, file_path)

        if file_name.endswith(".xlsx"):
            await message.answer(f"تم استلام ملف Excel: {file_name}. جاري معالجة البيانات...")
            try:
                process_excel_file(file_path)
                await message.answer("تمت معالجة ملف Excel بنجاح.")
            except Exception as e:
                await message.answer(f"حدث خطأ أثناء معالجة ملف Excel: {e}")
        elif file_name.endswith(".docx"):
            await message.answer(f"تم استلام ملف Word: {file_name}. جاري معالجة البيانات...")
            try:
                process_word_file(file_path)
                await message.answer("تمت معالجة ملف Word بنجاح.")
            except Exception as e:
                await message.answer(f"حدث خطأ أثناء معالجة ملف Word: {e}")
        else:
            await message.answer("صيغة الملف غير مدعومة. يرجى إرسال ملف Excel (.xlsx) أو Word (.docx).")

        await state.clear()

    # Admin Handlers
    @dp.message(F.text == "مشرف")
    async def cmd_admin(message: types.Message, state: FSMContext):
        await state.set_state(Admin.waiting_for_password)
        await message.answer("يرجى إدخال كلمة مرور المشرف:")

    @dp.message(Admin.waiting_for_password)
    async def process_admin_password(message: types.Message, state: FSMContext):
        if message.text == ADMIN_PASSWORD:
            await state.set_state(Admin.main_menu)
            keyboard = types.ReplyKeyboardMarkup(
                keyboard=[
                    [types.KeyboardButton(text="عرض إحصائيات الطلاب")],
                    [types.KeyboardButton(text="تصدير بيانات الطلاب")],
                    [types.KeyboardButton(text="إغلاق/فتح استمارة التقديم")],
                    [types.KeyboardButton(text="السماح/منع عرض بيانات الطلاب")],
                    [types.KeyboardButton(text="إدارة المشرفين")],
                    [types.KeyboardButton(text="العودة للقائمة الرئيسية")]
                ],
                resize_keyboard=True,
                one_time_keyboard=True
            )
            await message.answer("مرحباً بك أيها المشرف!", reply_markup=keyboard)
        else:
            await message.answer("كلمة المرور غير صحيحة. يرجى المحاولة مرة أخرى.")
            await state.clear()

    @dp.message(Admin.main_menu, F.text == "عرض إحصائيات الطلاب")
    async def show_student_statistics(message: types.Message, state: FSMContext):
        stats = get_student_statistics()
        response_message = "إحصائيات الطلاب:\n"
        response_message += f'العدد الكلي للطلاب: {stats.get("total_students", 0)}\n'
        response_message += "الطلاب حسب الصف:\n"
        for grade, count in stats.get("students_by_grade", {}).items():
            response_message += f"  {grade}: {count}\n"
        response_message += "الطلاب حسب الشعبة:\n"
        for section, count in stats.get("students_by_section", {}).items():
            response_message += f"  {section}: {count}\n"
        await message.answer(response_message)

    @dp.message(Admin.main_menu, F.text == "تصدير بيانات الطلاب")
    async def export_student_data(message: types.Message, state: FSMContext):
        file_name = "students_data.xlsx"
        file_path = os.path.join("exports", file_name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        if export_students_to_excel(file_path):
            await message.answer_document(types.FSInputFile(file_path), caption="تم تصدير بيانات الطلاب بنجاح.")
        else:
            await message.answer("لا توجد بيانات لتصديرها.")

    @dp.message(Admin.main_menu, F.text == "إغلاق/فتح استمارة التقديم")
    async def toggle_form_status(message: types.Message, state: FSMContext):
        current_status = get_setting("form_status")
        new_status = "closed" if current_status == "open" else "open"
        update_setting("form_status", new_status)
        await message.answer(f"تم {new_status} استمارة التقديم بنجاح.")

    @dp.message(Admin.main_menu, F.text == "السماح/منع عرض بيانات الطلاب")
    async def toggle_view_data_permission(message: types.Message, state: FSMContext):
        await state.set_state(Admin.toggle_view_data_name)
        await message.answer("يرجى إدخال الاسم الرباعي للطالب لتغيير صلاحية عرض البيانات:")

    @dp.message(Admin.toggle_view_data_name)
    async def process_toggle_view_data_name(message: types.Message, state: FSMContext):
        full_name = message.text.strip()
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, full_name, can_view_data FROM Students WHERE full_name = ?", (full_name,))
        student = cursor.fetchone()

        if student:
            student_id = student["id"]
            current_permission = student["can_view_data"]
            new_permission = 0 if current_permission == 1 else 1 # Toggle 0 (False) or 1 (True)
            permission_text = "السماح" if new_permission == 1 else "المنع"

            cursor.execute("UPDATE Students SET can_view_data = ? WHERE id = ?", (new_permission, student_id))
            conn.commit()
            await message.answer(f"تم {permission_text} للطالب {full_name} من عرض بياناته.")
        else:
            await message.answer(f"لم يتم العثور على الطالب {full_name}.")
        conn.close()
        await state.clear()

    @dp.message(Admin.main_menu, F.text == "إدارة المشرفين")
    async def manage_supervisors(message: types.Message, state: FSMContext):
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="إضافة مشرف")],
                [types.KeyboardButton(text="حذف مشرف")],
                [types.KeyboardButton(text="عرض المشرفين")],
                [types.KeyboardButton(text="العودة لقائمة المشرف")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer("خيارات إدارة المشرفين:", reply_markup=keyboard)
        await state.set_state(Admin.supervisor_management)

    @dp.message(Admin.supervisor_management, F.text == "إضافة مشرف")
    async def add_supervisor_start(message: types.Message, state: FSMContext):
        await state.set_state(Admin.add_supervisor_telegram_id)
        await message.answer("يرجى إدخال Telegram ID للمشرف الجديد:")

    @dp.message(Admin.add_supervisor_telegram_id)
    async def process_add_supervisor_telegram_id(message: types.Message, state: FSMContext):
        try:
            telegram_id = int(message.text)
            await state.update_data(new_supervisor_telegram_id=telegram_id)
            await state.set_state(Admin.add_supervisor_username)
            await message.answer("يرجى إدخال اسم المستخدم (username) للمشرف الجديد (اختياري، اكتب \"لا يوجد\" إذا لم يكن هناك):")
        except ValueError:
            await message.answer("Telegram ID غير صحيح. يرجى إدخال رقم صحيح.")

    @dp.message(Admin.add_supervisor_username)
    async def process_add_supervisor_username(message: types.Message, state: FSMContext):
        username = message.text.strip()
        if username == 'لا يوجد':
            username = None
        await state.update_data(new_supervisor_username=username)
        await state.set_state(Admin.add_supervisor_full_name)
        await message.answer("يرجى إدخال الاسم الكامل للمشرف الجديد:")

    @dp.message(Admin.add_supervisor_full_name)
    async def process_add_supervisor_full_name(message: types.Message, state: FSMContext):
        full_name = message.text.strip()
        await state.update_data(new_supervisor_full_name=full_name)
        await state.set_state(Admin.add_supervisor_password)
        await message.answer("يرجى إدخال كلمة المرور للمشرف الجديد:")

    @dp.message(Admin.add_supervisor_password)
    async def process_add_supervisor_password(message: types.Message, state: FSMContext):
        password = message.text.strip()
        user_data = await state.get_data()
        telegram_id = user_data.get("new_supervisor_telegram_id")
        username = user_data.get("new_supervisor_username")
        full_name = user_data.get("new_supervisor_full_name")

        if add_supervisor(telegram_id, username, full_name, password):
            await message.answer(f"تم إضافة المشرف {full_name} بنجاح.")
        else:
            await message.answer("حدث خطأ أثناء إضافة المشرف. قد يكون Telegram ID مستخدمًا بالفعل.")
        await state.clear()
        await manage_supervisors(message, state) # Return to supervisor management menu

    @dp.message(Admin.supervisor_management, F.text == "حذف مشرف")
    async def remove_supervisor_start(message: types.Message, state: FSMContext):
        await state.set_state(Admin.remove_supervisor_telegram_id)
        await message.answer("يرجى إدخال Telegram ID للمشرف الذي تود حذفه:")

    @dp.message(Admin.remove_supervisor_telegram_id)
    async def process_remove_supervisor_telegram_id(message: types.Message, state: FSMContext):
        try:
            telegram_id = int(message.text)
            if remove_supervisor(telegram_id):
                await message.answer(f"تم حذف المشرف ذو Telegram ID: {telegram_id} بنجاح.")
            else:
                await message.answer("لم يتم العثور على مشرف بهذا Telegram ID.")
        except ValueError:
            await message.answer("Telegram ID غير صحيح. يرجى إدخال رقم صحيح.")
        await state.clear()
        await manage_supervisors(message, state) # Return to supervisor management menu

    @dp.message(Admin.supervisor_management, F.text == "عرض المشرفين")
    async def view_supervisors(message: types.Message, state: FSMContext):
        supervisors = get_all_supervisors()
        if supervisors:
            response_message = "قائمة المشرفين:\n"
            for sup in supervisors:
                response_message += f"- الاسم: {sup['full_name']}, اسم المستخدم: {sup['username'] if sup['username'] else 'لا يوجد'}, Telegram ID: {sup['telegram_id']}\n"
            await message.answer(response_message)
        else:
            await message.answer("لا يوجد مشرفون مسجلون حالياً.")
        await state.clear()
        await manage_supervisors(message, state) # Return to supervisor management menu

    @dp.message(Admin.supervisor_management, F.text == "العودة لقائمة المشرف")
    async def back_to_admin_menu(message: types.Message, state: FSMContext):
        await state.set_state(Admin.main_menu)
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text="عرض إحصائيات الطلاب")],
                [types.KeyboardButton(text="تصدير بيانات الطلاب")],
                [types.KeyboardButton(text="إغلاق/فتح استمارة التقديم")],
                [types.KeyboardButton(text="السماح/منع عرض بيانات الطلاب")],
                [types.KeyboardButton(text="إدارة المشرفين")],
                [types.KeyboardButton(text="العودة للقائمة الرئيسية")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer("تم العودة إلى قائمة المشرف.", reply_markup=keyboard)

    @dp.message(Admin.main_menu, F.text == "العودة للقائمة الرئيسية")
    async def back_to_main_menu(message: types.Message, state: FSMContext):
        await state.clear()
        await command_start_handler(message)

    # Start polling
    await dp.start_polling(bot)

if __name__ == "__main__":
    import asyncio
    asyncio.run(main())




